from __future__ import annotations

import importlib
import json
from pathlib import Path
from typing import TYPE_CHECKING, Any

import numpy as np
from himena.types import WidgetDataModel
from himena.standards.model_meta import TableMeta, TextMeta
from himena.consts import StandardType

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet
    from himena.standards import plotting as hplt


def default_text_reader(file_path: Path) -> WidgetDataModel:
    """Read text file."""
    import chardet

    if file_path.suffix in (".html", ".htm"):
        typ = StandardType.HTML
    elif file_path.suffix == ".json":
        typ = StandardType.JSON
    elif file_path.suffix == ".svg":
        typ = StandardType.SVG
    elif file_path.suffix == ".ipynb":
        typ = StandardType.IPYNB
    else:
        typ = StandardType.TEXT
    with file_path.open("rb") as f:
        detector = chardet.UniversalDetector()
        for line in f:
            detector.feed(line)
            if detector.done:
                break
        detector.close()
    encoding = detector.result["encoding"]
    value = file_path.read_text(encoding=encoding)
    return WidgetDataModel(
        value=value,
        type=typ,
        source=file_path,
        extension_default=file_path.suffix,
        metadata=TextMeta(encoding=encoding),
    )


def default_image_reader(file_path: Path) -> WidgetDataModel:
    """Read image file."""
    from PIL import Image

    arr = np.array(Image.open(file_path))

    return WidgetDataModel(
        value=arr,
        type=StandardType.IMAGE,
        extension_default=file_path.suffix,
    )


def _read_txt_as_numpy(file_path: Path, delimiter: str):
    arr = np.loadtxt(
        file_path,
        dtype=np.dtypes.StringDType(),
        delimiter=delimiter,
    )
    return WidgetDataModel(
        value=arr,
        type=StandardType.TABLE,
        extension_default=file_path.suffix,
        metadata=TableMeta(separator=delimiter),
    )


def default_csv_reader(file_path: Path) -> WidgetDataModel:
    """Read CSV file."""
    return _read_txt_as_numpy(file_path, ",")


def default_tsv_reader(file_path: Path) -> WidgetDataModel:
    """Read TSV file."""
    return _read_txt_as_numpy(file_path, "\t")


def default_plot_reader(file_path: Path) -> WidgetDataModel:
    """Write plot layout to a json file."""
    from himena.standards import plotting

    with open(file_path) as f:
        js = json.load(f)
        if not isinstance(js, dict):
            raise ValueError(f"Expected a dictionary, got {type(js)}.")
        if not (typ := js.pop("type")):
            raise ValueError("'type' field not found in the JSON file.")
        plot_layout = plotting.BaseLayoutModel.construct(typ, js)
    return WidgetDataModel(
        value=plot_layout,
        type=StandardType.PLOT,
        extension_default=".plot.json",
    )


def default_excel_reader(file_path: Path) -> WidgetDataModel:
    """Read Excel file."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=False)
    data = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_data = []
        for row in ws.iter_rows():
            row_input = []
            for cell in row:
                if cell.value is None:
                    row_input.append("")
                else:
                    row_input.append(str(cell.value))
            sheet_data.append(row_input)

        data[sheet] = sheet_data

    return WidgetDataModel(
        value=data,
        type=StandardType.EXCEL,
        extension_default=file_path.suffix,
    )


def default_array_reader(file_path: Path) -> WidgetDataModel:
    """Read array file."""
    arr = np.load(file_path)
    return WidgetDataModel(value=arr, type=StandardType.ARRAY)


def default_pickle_reader(file_path: Path) -> WidgetDataModel:
    """Read pickle file."""
    import pickle

    with file_path.open("rb") as f:
        value = pickle.load(f)
    if isinstance(value, WidgetDataModel):
        # picke is created by himena.
        return value
    else:
        # pickle is created probably by other library. Just read as type "any".
        return WidgetDataModel(value=value, type=StandardType.ANY)


def fallback_reader(file_path: Path | list[Path]) -> WidgetDataModel:
    return WidgetDataModel(value=file_path, type=StandardType.READER_NOT_FOUND)


class DataFrameReader:
    def __init__(self, module: str, method: str, kwargs: dict[str, Any]):
        self._module = module
        self._method = method
        self._kwargs = kwargs

    def __repr__(self):
        return f"{self.__class__.__name__}<{self._module}.{self._method}>"

    def __call__(self, file_path: Path) -> WidgetDataModel:
        mod = importlib.import_module(self._module)
        method = getattr(mod, self._method)
        df = method(file_path, **self._kwargs)
        return WidgetDataModel(value=df, type=StandardType.DATAFRAME)


def default_text_writer(model: WidgetDataModel[str], path: Path) -> None:
    """Write text file."""
    if isinstance(meta := model.metadata, TextMeta):
        encoding = meta.encoding
    return path.write_text(model.value, encoding=encoding)


def default_table_writer(model: WidgetDataModel[np.ndarray], path: Path) -> None:
    """Write table data to a text file."""
    delimiter = None
    if isinstance(meta := model.metadata, TableMeta):
        delimiter = meta.separator
    if delimiter is None:
        if path.suffix == ".tsv":
            delimiter = "\t"
        else:
            delimiter = ","
    np.savetxt(path, model.value, fmt="%s", delimiter=delimiter)


def default_image_writer(model: WidgetDataModel[np.ndarray], path: Path) -> None:
    """Write image file."""
    from PIL import Image

    Image.fromarray(model.value).save(path)
    return None


def default_dict_writer(model: WidgetDataModel[dict[str, Any]], path: Path) -> None:
    """Write parameters to a json file."""
    with path.open("w") as f:
        json.dump(model.value, f, default=_json_default)
    return None


def default_plot_writer(
    model: WidgetDataModel[hplt.BaseLayoutModel], path: Path
) -> None:
    """Write plot layout to a json file."""
    js = model.value.model_dump_typed()
    with path.open("w") as f:
        json.dump(js, f, default=_json_default)
    return None


def default_excel_writer(
    model: WidgetDataModel[dict[str, np.ndarray]],
    path: Path,
) -> None:
    """Write Excel file."""
    import openpyxl

    wb = openpyxl.Workbook()
    if active_sheet := wb.active:
        wb.remove(active_sheet)
    for sheet_name, table in model.value.items():
        ws: Worksheet = wb.create_sheet(sheet_name)
        for r, row in enumerate(table):
            for c, cell_str in enumerate(row):
                cell_str: str
                if cell_str.startswith("="):
                    cell_data_type = "f"
                else:
                    try:
                        float(cell_str)
                        cell_data_type = "n"
                    except ValueError:
                        cell_data_type = "s"
                ws.cell(r + 1, c + 1).value = cell_str
                ws.cell(r + 1, c + 1).data_type = cell_data_type
    wb.save(path)
    return None


def default_array_writer(
    model: WidgetDataModel[np.ndarray],
    path: Path,
) -> None:
    """Write array file."""
    np.save(path, model.value)
    return None


def default_dataframe_writer(
    model: WidgetDataModel[dict[str, np.ndarray]],
    path: Path,
) -> None:
    """Write dataframe file."""
    from himena._data_wrappers import wrap_dataframe

    return wrap_dataframe(model.value).write(path)


def default_pickle_writer(
    model: WidgetDataModel[Any],
    path: Path,
) -> None:
    """Write pickle file."""
    import pickle

    with path.open("wb") as f:
        pickle.dump(model.value, f)
    return None


def _json_default(obj):
    import cmap

    if isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, cmap.Color):
        return obj.hex
    elif isinstance(obj, cmap.Colormap):
        return obj.name
    raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable.")
