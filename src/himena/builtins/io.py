from __future__ import annotations

import importlib
from pathlib import Path
from typing import TYPE_CHECKING, Any

import numpy as np
from himena.plugins import register_reader_provider, register_writer_provider
from himena.types import WidgetDataModel
from himena.model_meta import TextMeta
from himena.consts import (
    StandardType,
    BasicTextFileTypes,
    ConventionalTextFileNames,
    ExcelFileTypes,
)

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


def default_text_reader(file_path: Path) -> WidgetDataModel:
    """Read text file."""
    return _read_text(file_path, StandardType.TEXT)


def default_html_reader(file_path: Path) -> WidgetDataModel:
    """Read HTML file."""
    return _read_text(file_path, StandardType.HTML)


def _read_text(file: Path, typ: str) -> tuple[str, str]:
    """Read text file with auto-detected encoding."""
    import chardet

    with file.open("rb") as f:
        detector = chardet.UniversalDetector()
        for line in f:
            detector.feed(line)
            if detector.done:
                break
        detector.close()
    encoding = detector.result["encoding"]
    value = file.read_text(encoding=encoding)
    return WidgetDataModel(
        value=value,
        type=typ,
        source=file,
        metadata=TextMeta(encoding=encoding),
    )


def default_image_reader(file_path: Path) -> WidgetDataModel:
    """Read image file."""
    from PIL import Image

    arr = np.array(Image.open(file_path))

    return WidgetDataModel(
        value=arr,
        type=StandardType.IMAGE,
    )


def _read_txt_as_numpy(file_path: Path, delimiter: str):
    arr = np.loadtxt(
        file_path,
        dtype=np.dtypes.StringDType(),
        delimiter=delimiter,
    )
    return WidgetDataModel(
        value=arr,
        type=StandardType.TABLE,
        extension_default=file_path.suffix,
    )


def default_csv_reader(file_path: Path) -> WidgetDataModel:
    """Read CSV file."""
    return _read_txt_as_numpy(file_path, ",")


def default_tsv_reader(file_path: Path) -> WidgetDataModel:
    """Read TSV file."""
    return _read_txt_as_numpy(file_path, "\t")


def default_excel_reader(file_path: Path) -> WidgetDataModel:
    """Read Excel file."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=False)
    data = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_data = []
        for row in ws.iter_rows():
            row_input = []
            for cell in row:
                if cell.value is None:
                    row_input.append("")
                else:
                    row_input.append(str(cell.value))
            sheet_data.append(row_input)

        data[sheet] = sheet_data

    return WidgetDataModel(
        value=data,
        type=StandardType.EXCEL,
        extension_default=file_path.suffix,
    )


def default_array_reader(file_path: Path) -> WidgetDataModel:
    """Read array file."""
    arr = np.load(file_path)
    return WidgetDataModel(
        value=arr,
        type=StandardType.ARRAY,
    )


@register_reader_provider(priority=50)
def default_reader_provider(file_path: Path | list[Path]):
    """Get default reader."""
    if isinstance(file_path, list):
        return None
    if file_path.suffix in (".html", ".htm"):
        return default_html_reader
    elif file_path.suffix in BasicTextFileTypes:
        return default_text_reader
    elif file_path.suffix == ".csv":
        return default_csv_reader
    elif file_path.suffix == ".tsv":
        return default_tsv_reader
    elif file_path.suffix in {".png", ".jpg", ".jpeg"}:
        return default_image_reader
    elif file_path.name in ConventionalTextFileNames:
        return default_text_reader
    elif file_path.suffix in ExcelFileTypes:
        return default_excel_reader
    elif file_path.suffix == ".npy":
        return default_array_reader
    elif file_path.suffix in {".parquet", ".pq"}:
        return DataFrameReader("pandas", "read_parquet", {})
    return None


@register_reader_provider(priority=-100)
def read_as_text_anyway_provider(file_path: Path) -> WidgetDataModel:
    return default_text_reader


def fallback_reader(file_path: Path | list[Path]) -> WidgetDataModel:
    return WidgetDataModel(value=file_path, type=StandardType.READER_NOT_FOUND)


@register_reader_provider(priority=0)
def read_as_unknown_provider(file_path: Path) -> WidgetDataModel:
    return fallback_reader


class DataFrameReader:
    def __init__(self, module: str, method: str, kwargs: dict[str, Any]):
        self._module = module
        self._method = method
        self._kwargs = kwargs

    def __repr__(self):
        return f"{self.__class__.__name__}<{self._module}.{self._method}>"

    def __call__(self, file_path: Path) -> WidgetDataModel:
        mod = importlib.import_module(self._module)
        method = getattr(mod, self._method)
        df = method(file_path, **self._kwargs)
        return WidgetDataModel(value=df, type=StandardType.DATAFRAME)


@register_reader_provider(priority=-50)
def pandas_reader_provider(file_path: Path) -> WidgetDataModel:
    """Read dataframe using pandas."""
    if file_path.suffix in (".html", ".htm"):
        _reader = "pandas", "read_html", {}
    elif file_path.suffix in (".csv", ".txt"):
        _reader = "pandas", "read_csv", {}
    elif file_path.suffix == ".tsv":
        _reader = "pandas", "read_csv", {"sep": "\t"}
    elif file_path.suffix == ".json":
        _reader = "pandas", "read_json", {}
    elif file_path.suffix in (".pq", ".parquet"):
        _reader = "pandas", "read_parquet", {}
    elif file_path.suffix == ".feather":
        _reader = "pandas", "read_feather", {}
    else:
        return None
    return DataFrameReader(*_reader)


@register_reader_provider(priority=-50)
def polars_reader_provider(file_path: Path) -> WidgetDataModel:
    """Read dataframe using polars."""
    if isinstance(file_path, list):
        return None
    elif file_path.suffix in (".csv", ".txt"):
        _reader = "polars", "read_csv", {}
    elif file_path.suffix == ".tsv":
        _reader = "polars", "read_csv", {"sep": "\t"}
    elif file_path.suffix == ".feather":
        _reader = "polars", "read_ipc", {}
    elif file_path.suffix == ".json":
        _reader = "polars", "read_json", {}
    elif file_path.suffix in (".parquet", ".pq"):
        _reader = "polars", "read_parquet", {}
    else:
        return None
    return DataFrameReader(*_reader)


def default_text_writer(model: WidgetDataModel[str], path: Path) -> None:
    """Write text file."""
    if isinstance(meta := model.metadata, TextMeta):
        encoding = meta.encoding
    return path.write_text(model.value, encoding=encoding)


def default_csv_writer(model: WidgetDataModel[np.ndarray], path: Path) -> None:
    """Write CSV file."""
    if path.suffix == ".tsv":
        delimiter = "\t"
    else:
        delimiter = ","
    np.savetxt(path, model.value, fmt="%s", delimiter=delimiter)


def default_image_writer(model: WidgetDataModel[np.ndarray], path: Path) -> None:
    """Write image file."""
    from PIL import Image

    Image.fromarray(model.value).save(path)
    return None


def default_parameter_writer(
    model: WidgetDataModel[dict[str, Any]], path: Path
) -> None:
    """Write parameters to a json file."""
    import json

    with open(path, "w") as f:
        json.dump(model.value, f)


def default_excel_writer(
    model: WidgetDataModel[dict[str, np.ndarray]],
    path: Path,
) -> None:
    """Write Excel file."""
    import openpyxl

    wb = openpyxl.Workbook()
    if active_sheet := wb.active:
        wb.remove(active_sheet)
    for sheet_name, table in model.value.items():
        ws: Worksheet = wb.create_sheet(sheet_name)
        for r, row in enumerate(table):
            for c, cell_str in enumerate(row):
                cell_str: str
                if cell_str.startswith("="):
                    cell_data_type = "f"
                else:
                    try:
                        float(cell_str)
                        cell_data_type = "n"
                    except ValueError:
                        cell_data_type = "s"
                ws.cell(r + 1, c + 1).value = cell_str
                ws.cell(r + 1, c + 1).data_type = cell_data_type
    wb.save(path)
    return None


def default_array_writer(
    model: WidgetDataModel[np.ndarray],
    path: Path,
) -> None:
    """Write array file."""
    np.save(path, model.value)
    return None


def default_dataframe_writer(
    model: WidgetDataModel[dict[str, np.ndarray]],
    path: Path,
) -> None:
    """Write dataframe file."""
    from himena._data_wrappers import wrap_dataframe

    return wrap_dataframe(model.value).write(path)


@register_writer_provider(priority=50)
def default_writer_provider(model: WidgetDataModel):
    """Get default writer."""
    if model.type is None:
        return None
    if model.is_subtype_of(StandardType.TEXT):
        return default_text_writer
    elif model.is_subtype_of(StandardType.TABLE):
        return default_csv_writer
    elif model.is_subtype_of(StandardType.IMAGE):
        return default_image_writer
    elif model.is_subtype_of(StandardType.PARAMETERS):
        return default_parameter_writer
    elif model.is_subtype_of(StandardType.EXCEL):
        return default_excel_writer
    elif model.is_subtype_of(StandardType.ARRAY):
        return default_array_writer
    elif model.is_subtype_of(StandardType.DATAFRAME):
        return default_dataframe_writer
    else:
        return None
